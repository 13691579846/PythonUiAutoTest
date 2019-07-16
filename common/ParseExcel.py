"""
------------------------------------
@Time : 2019/7/13 19:57
@Auth : linux超
@File : ParseExcel.py
@IDE  : PyCharm
@Motto: Real warriors,dare to face the bleak warning,dare to face the incisive error!
@QQ   : 28174043@qq.com
@GROUP: 878565760
------------------------------------
"""
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.styles.colors import (BLACK, RED, GREEN)
from collections import namedtuple

from config.config import DATA_PATH


class ParseExcel(object):

    def __init__(self, filename):
        self.filename = filename
        self.wb = load_workbook(self.filename)
        self.font = Font(color=None)
        self.RGBDict = {'red': RED, 'green': GREEN, 'black': BLACK}

    def max_row(self, sheet_name):
        return self.wb[sheet_name].max_row

    def min_row(self, sheet_name):
        return self.wb[sheet_name].min_row

    def max_column(self, sheet_name):
        return self.wb[sheet_name].max_column

    def min_column(self, sheet_name):
        return self.wb[sheet_name].min_column

    def head(self, sheet_name):
        min_row = self.min_row(sheet_name)
        head = tuple(self.wb[sheet_name].iter_rows(max_row=min_row, values_only=True))[0]
        return head

    def all_values_tuple(self, sheet_name):
        """get name tuple of all values for excel"""
        case_list = []
        head = self.head(sheet_name)
        case = namedtuple('case', head)
        all_values = self.wb[sheet_name].iter_rows(min_row=self.min_row(sheet_name) + 1,
                                                   max_col=self.max_column(sheet_name),
                                                   values_only=True)
        for value in all_values:
            case_list.append(case(*value))
        return case_list

    def all_value_dict(self, sheet_name):
        head = self.head(sheet_name)
        values = self.wb[sheet_name].iter_rows(min_row=self.min_row(sheet_name) + 1,
                                               max_col=self.max_column(sheet_name),
                                               values_only=True)
        value_list = []
        for value in values:
            value_list.append(dict(zip(head, value)))
        return value_list

    def write_cell(self, sheet_name, row, column, value, color='black'):
        """write cell value with color"""
        if isinstance(row, int) and isinstance(column, int):
            try:
                cell_obj = self.wb[sheet_name].cell(row, column)
                cell_obj.font = Font(color=self.RGBDict[color], bold=True)
                cell_obj.value = value
                self.wb.save(self.filename)
            except Exception as e:
                raise e
        else:
            raise TypeError('row and column must be type int')

    def __call__(self, sheet_name):
        return self.all_value_dict(sheet_name)


do_excel = ParseExcel(DATA_PATH)


if __name__ == '__main__':
    excel = ParseExcel(r'D:\PythonUiAutoTest\datas\test_cases.xlsx')
    print(excel.all_value_dict('login'))
